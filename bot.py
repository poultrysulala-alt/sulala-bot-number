#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import re, zipfile, json, logging
from pathlib import Path
from datetime import datetime
from collections import Counter
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes

BOT_TOKEN  = "8792112088:AAHPOQKqDcz2ULmm-uH3o4GgG73iAPGMCP4"
ADMIN_ID   = 45138916
BASE_DIR   = Path(__file__).parent
EXCEL      = BASE_DIR / "ارقام_الدواجن___شركات_-_افراد__.xlsx"
USERS_FILE = BASE_DIR / "users.json"
TEMP_DIR   = BASE_DIR / "temp"
TEMP_DIR.mkdir(exist_ok=True)
logging.basicConfig(level=logging.WARNING)

FILL_ROW  = PatternFill("solid", fgColor="F1F8E9")
FONT_MAIN = Font(name="Arial", size=11, color="212121")
ALIGN_C   = Alignment(horizontal="center", vertical="center")
ALIGN_R   = Alignment(horizontal="right",  vertical="center")

# ─── المستخدمون ───────────────────────────────────────────────────────────────
def get_users():
    if USERS_FILE.exists():
        return json.loads(USERS_FILE.read_text(encoding="utf-8"))
    return [ADMIN_ID]

def set_users(u):
    USERS_FILE.write_text(json.dumps(u), encoding="utf-8")

def ok(uid):
    return uid in get_users()

# ─── كلمات الشركات والمناطق ───────────────────────────────────────────────────
COMPANY_KW = ["شركة","مؤسسة","مصنع","توزيع","تسويق","مندوب","مبيعات",
               "وكيل","موزع","مشرف","للتجارة","المحدودة","شركه","مؤسسه"]

REGIONS = ["الرياض","جدة","مكة","المدينة","الدمام","الخبر","الخرج",
           "القصيم","بريدة","حائل","تبوك","أبها","خميس مشيط","نجران",
           "جيزان","الطائف","ينبع","الأحساء","القطيف","الجبيل","عنيزة"]

def get_phone(text):
    # أرقام سعودية فقط - تبدأ بـ 05
    m = re.search(r'05\d{8}', text)
    if m:
        return m.group()
    # أرقام بصيغة دولية سعودية
    m = re.search(r'(?:\+966|00966)5\d{8}', text)
    if m:
        n = re.sub(r'\D','', m.group())
        n = re.sub(r'^00966|^966','', n)
        return '0' + n
    return None

def get_name(text):
    m = re.search(r'~\s*([^:]+):', text)
    if m: return m.group(1).strip()[:40]
    m = re.search(r'[أا]بو\s+\w+', text)
    if m: return m.group().strip()
    return "غير محدد"

def get_region(text):
    for r in REGIONS:
        if r in text: return r
    return "غير محدد"

def is_co(text):
    return any(k in text for k in COMPANY_KW)

def clean_source(source):
    # تنظيف اسم المصدر
    src = source
    if '_WhatsApp' in src:
        src = src.split('_WhatsApp')[0]
    if 'WhatsApp_Chat_' in src:
        src = src.replace('WhatsApp_Chat_', '')
    return src[:25]

# ─── قراءة ملف واتساب ────────────────────────────────────────────────────────
def read_chat(path):
    try: txt = Path(path).read_text(encoding='utf-8')
    except: txt = Path(path).read_text(encoding='utf-8-sig')
    out = []
    for line in txt.splitlines():
        p = get_phone(line)
        if p and len(line) > 15:
            date_m = re.search(r'(\d{1,2}[/.]\d{1,2}[/.]\d{2,4})', line)
            date = date_m.group(1) if date_m else datetime.now().strftime('%d.%m.%Y')
            out.append({'phone': p, 'text': line, 'date': date})
    return out

def write_cell(ws, row, col, value, align="right"):
    cell = ws.cell(row, col, value)
    cell.fill = FILL_ROW
    cell.font = FONT_MAIN
    cell.alignment = ALIGN_C if align == "center" else ALIGN_R
    return cell

def update_summary(wb, total_a, total_s):
    ws = wb['الملخص']
    ws['C5'] = total_a
    ws['C6'] = total_s
    ws['C7'] = total_a + total_s

# ─── تحديث الإكسيل ────────────────────────────────────────────────────────────
def update_excel(records, source):
    df_a = pd.read_excel(EXCEL, sheet_name='أفراد',  header=1)
    df_s = pd.read_excel(EXCEL, sheet_name='شركات', header=1)

    exist = set()
    if 'رقم_التواصل' in df_a.columns:
        exist.update(str(x)[-9:] for x in df_a['رقم_التواصل'].dropna())
    if 'رقم الجوال' in df_s.columns:
        exist.update(str(x)[-9:] for x in df_s['رقم الجوال'].dropna())

    src = clean_source(source)
    new_a, new_s, dups = [], [], 0

    for r in records:
        ph = r['phone']
        if ph[-9:] in exist:
            dups += 1
            continue
        exist.add(ph[-9:])
        if is_co(r['text']):
            new_s.append((get_name(r['text']), ph, src))
        else:
            new_a.append((get_name(r['text']), ph, get_region(r['text']), r['date'], r['text'][:200], src))

    wb = load_workbook(EXCEL)

    if new_a:
        ws = wb['أفراد']
        start = ws.max_row + 1
        num   = len(df_a) + 1
        for i, (nm, ph, rg, dt, tx, s) in enumerate(new_a):
            r = start + i
            write_cell(ws, r, 1, num+i, "center")
            write_cell(ws, r, 2, ph,    "center")
            write_cell(ws, r, 3, nm,    "right")
            write_cell(ws, r, 4, 'دواجن', "right")
            write_cell(ws, r, 5, rg,    "right")
            write_cell(ws, r, 6, '',    "center")
            write_cell(ws, r, 7, dt,    "center")
            write_cell(ws, r, 8, tx,    "right")
            write_cell(ws, r, 9, s,     "right")

    if new_s:
        ws = wb['شركات']
        start = ws.max_row + 1
        num   = len(df_s) + 1
        for i, (nm, ph, s) in enumerate(new_s):
            r = start + i
            write_cell(ws, r, 1, num+i,      "center")
            write_cell(ws, r, 2, nm,         "right")
            write_cell(ws, r, 3, 'غير محدد', "right")
            write_cell(ws, r, 4, ph,         "center")

    total_a = len(df_a) + len(new_a)
    total_s = len(df_s) + len(new_s)
    update_summary(wb, total_a, total_s)
    wb.save(EXCEL)
    return len(new_a), len(new_s), dups, total_a + total_s

# ─── معالجة الملف ─────────────────────────────────────────────────────────────
def process(file_path):
    fp = Path(file_path)
    source = fp.stem
    recs = []

    if fp.suffix.lower() == '.zip':
        out = TEMP_DIR / fp.stem
        out.mkdir(exist_ok=True)
        with zipfile.ZipFile(fp) as z:
            z.extractall(out)
        for t in out.rglob("*.txt"):
            recs.extend(read_chat(t))
        import shutil; shutil.rmtree(out, ignore_errors=True)
    elif fp.suffix.lower() == '.txt':
        recs = read_chat(fp)

    fp.unlink(missing_ok=True)

    if not recs:
        return None, "⚠️ ما وجدت أرقام سعودية في الملف"

    na, ns, dups, total = update_excel(recs, source)
    msg = (f"✅ *تم بنجاح!*\n\n"
           f"📁 المصدر: `{clean_source(source)}`\n"
           f"👤 أفراد جدد: *{na}*\n"
           f"🏢 شركات جديدة: *{ns}*\n"
           f"🔁 مكررة: *{dups}*\n"
           f"📊 الإجمالي: *{total:,}* رقم\n"
           f"🕐 {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    return (na, ns), msg

# ─── كشف المكررة داخل الملف ──────────────────────────────────────────────────
def find_duplicates():
    df_a = pd.read_excel(EXCEL, sheet_name='أفراد',  header=1)
    df_s = pd.read_excel(EXCEL, sheet_name='شركات', header=1)

    phones_a = df_a['رقم_التواصل'].dropna().astype(str).tolist() if 'رقم_التواصل' in df_a.columns else []
    phones_s = df_s['رقم الجوال'].dropna().astype(str).tolist()  if 'رقم الجوال'  in df_s.columns else []
    all_phones = phones_a + phones_s

    count = Counter(str(p)[-9:] for p in all_phones)
    dups = {p: c for p, c in count.items() if c > 1}
    return dups

# ─── أوامر البوت ──────────────────────────────────────────────────────────────
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not ok(update.effective_user.id):
        await update.message.reply_text("⛔ غير مصرح"); return
    await update.message.reply_text(
        "🐔 *بوت دواجن*\n\n"
        "أرسل ZIP أو TXT تصدير واتساب!\n\n"
        "الأوامر:\n"
        "/stats — إحصائيات تفصيلية\n"
        "/getfile — استلام ملف الإكسيل\n"
        "/duplicates — كشف الأرقام المكررة\n"
        "/adduser ID — إضافة مستخدم\n"
        "/users — قائمة المستخدمين",
        parse_mode="Markdown")

async def cmd_stats(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not ok(update.effective_user.id): return
    try:
        df_a = pd.read_excel(EXCEL, sheet_name='أفراد', header=1)
        df_s = pd.read_excel(EXCEL, sheet_name='شركات', header=1)

        # إحصائيات المناطق
        regions_txt = ""
        if 'المنطقة' in df_a.columns:
            regions = df_a['المنطقة'].value_counts().head(5)
            for rg, cnt in regions.items():
                if rg and str(rg) != 'nan':
                    regions_txt += f"  • {rg}: {cnt}\n"

        msg = (f"📊 *إحصائيات تفصيلية*\n\n"
               f"👤 أفراد: *{len(df_a):,}*\n"
               f"🏢 شركات: *{len(df_s):,}*\n"
               f"📱 الإجمالي: *{len(df_a)+len(df_s):,}*\n\n"
               f"🗺 *أكثر المناطق:*\n{regions_txt}")
        await update.message.reply_text(msg, parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"خطأ: {e}")

async def cmd_getfile(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not ok(update.effective_user.id):
        await update.message.reply_text("⛔ غير مصرح"); return
    try:
        await update.message.reply_document(
            document=open(EXCEL, 'rb'),
            filename="ارقام_الدواجن.xlsx",
            caption=f"📊 ملف الإكسيل المحدث\n🕐 {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    except Exception as e:
        await update.message.reply_text(f"❌ خطأ: {e}")

async def cmd_duplicates(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not ok(update.effective_user.id): return
    msg = await update.message.reply_text("⏳ جاري الفحص...")
    try:
        dups = find_duplicates()
        if not dups:
            await msg.edit_text("✅ لا توجد أرقام مكررة في قاعدة البيانات")
        else:
            txt = f"⚠️ *وجدت {len(dups)} رقم مكرر:*\n\n"
            for p, c in list(dups.items())[:20]:
                txt += f"• `{p}` — تكرر {c} مرات\n"
            if len(dups) > 20:
                txt += f"\n_وأكثر..._"
            await msg.edit_text(txt, parse_mode="Markdown")
    except Exception as e:
        await msg.edit_text(f"❌ خطأ: {e}")

async def cmd_adduser(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ للمشرف فقط"); return
    if not ctx.args:
        await update.message.reply_text("الاستخدام: /adduser 123456789"); return
    users = get_users()
    uid = int(ctx.args[0])
    if uid not in users:
        users.append(uid); set_users(users)
        await update.message.reply_text(f"✅ تمت إضافة {uid}")
    else:
        await update.message.reply_text("موجود مسبقاً")

async def cmd_users(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID: return
    u = get_users()
    txt = "\n".join([f"• {x}" + (" 👑" if x==ADMIN_ID else "") for x in u])
    await update.message.reply_text(f"👥 *المستخدمون:*\n{txt}", parse_mode="Markdown")

async def handle_doc(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not ok(update.effective_user.id):
        await update.message.reply_text("⛔ غير مصرح"); return
    doc = update.message.document
    if not (doc.file_name.lower().endswith('.zip') or doc.file_name.lower().endswith('.txt')):
        await update.message.reply_text("⚠️ أرسل ZIP أو TXT فقط"); return

    sender = update.effective_user.first_name or "مستخدم"
    msg = await update.message.reply_text("⏳ جاري المعالجة...")
    try:
        f  = await ctx.bot.get_file(doc.file_id)
        fp = TEMP_DIR / doc.file_name
        await f.download_to_drive(fp)
        result, text = process(fp)
        await msg.edit_text(text, parse_mode="Markdown")

        # إشعار للمشرف لو المرسل شخص ثاني
        if result and update.effective_user.id != ADMIN_ID:
            na, ns = result
            await ctx.bot.send_message(
                ADMIN_ID,
                f"🔔 *إشعار جديد*\n\n"
                f"👤 {sender} أضاف:\n"
                f"• أفراد: *{na}*\n"
                f"• شركات: *{ns}*",
                parse_mode="Markdown")
    except Exception as e:
        await msg.edit_text(f"❌ خطأ: {e}")

# ─── التشغيل ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("🐔 بوت الدواجن يعمل...")
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start",      cmd_start))
    app.add_handler(CommandHandler("stats",      cmd_stats))
    app.add_handler(CommandHandler("getfile",    cmd_getfile))
    app.add_handler(CommandHandler("duplicates", cmd_duplicates))
    app.add_handler(CommandHandler("adduser",    cmd_adduser))
    app.add_handler(CommandHandler("users",      cmd_users))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_doc))
    print("✅ البوت جاهز — افتح تيليقرام وأرسل /start")
    app.run_polling(drop_pending_updates=True)
